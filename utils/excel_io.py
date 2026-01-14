# utils/excel_io.py
from __future__ import annotations
import io
from typing import List, Tuple, Optional
import pandas as pd
import openpyxl
from utils.rules import get_subject_abbr

NAME_KEYS = ["họ tên", "họ và tên", "tên học sinh"]
LEVEL_KEYS = ["mức đạt được", "mức đạt", "xếp loại"]
SCORE_KEYS = ["điểm ktđk", "ktđk", "điểm kiểm tra"]

def _find_header_row(ws, max_scan_rows=15) -> Tuple[int, int, int, int]:
    """
    trả về: (header_row_index_1based, name_col, level_col, score_col) - col là 0-based
    """
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            row_vals.append(str(v).strip().lower() if v is not None else "")
        def find_idx(keys):
            for i, cell in enumerate(row_vals):
                if any(k in cell for k in keys):
                    return i
            return -1

        name_idx = find_idx(NAME_KEYS)
        if name_idx != -1:
            level_idx = find_idx(LEVEL_KEYS)
            score_idx = find_idx(SCORE_KEYS)
            return r, name_idx, level_idx, score_idx
    return -1, -1, -1, -1

def read_students_from_excel(file_bytes: bytes, selected_subject: str) -> List[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    subject_lower = selected_subject.lower()
    subject_abbr = get_subject_abbr(selected_subject).lower()

    # tìm sheet phù hợp (giống logic của bạn)
    sheet_name = None
    for n in wb.sheetnames:
        nl = n.lower()
        if subject_lower in nl or nl == subject_abbr or subject_abbr in nl:
            sheet_name = n
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    header_row, name_col, level_col, score_col = _find_header_row(ws)
    if name_col == -1:
        raise ValueError(f'Không tìm thấy cột "Họ Tên" trong sheet "{sheet_name}".')

    records: List[dict] = []
    stt = 0
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, name_col + 1).value
        name = str(name).strip() if name is not None else ""
        if len(name) <= 1:
            continue

        raw_level = ws.cell(r, level_col + 1).value if level_col != -1 else ""
        raw_level = str(raw_level).strip().upper() if raw_level is not None else ""
        level = ""
        if "TỐT" in raw_level or raw_level in ["T"] or "HTT" in raw_level:
            level = "T"
        elif "CHƯA" in raw_level or raw_level in ["C"] or "CHT" in raw_level:
            level = "C"
        elif "HOÀN THÀNH" in raw_level or raw_level in ["H", "HT"]:
            level = "H"

        raw_score = ws.cell(r, score_col + 1).value if score_col != -1 else 0
        try:
            diem = float(str(raw_score).replace(",", ".")) if raw_score not in [None, ""] else 0.0
        except Exception:
            diem = 0.0

        stt += 1
        records.append({
            "stt": stt,
            "hoTen": name,
            "ngaySinh": "",
            "diem": diem,
            "mucDo": level,
            "maNhanXet": "",
            "noiDung": "",
        })
    return records

def export_students_excel(records: List[dict]) -> bytes:
    df = pd.DataFrame([{
        "STT": r["stt"],
        "Họ tên": r["hoTen"],
        "Mức đạt được": "HTT" if r["mucDo"] == "T" else ("HT" if r["mucDo"] == "H" else "CHT"),
        "Điểm KTĐK": r["diem"] if (r.get("diem", 0) or 0) > 0 else "",
        "Mã NX": r["maNhanXet"],
        "Nội dung nhận xét": r["noiDung"],
    } for r in records])

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="NhanXet")
    return out.getvalue()

def export_bank_excel(bank: List[dict]) -> bytes:
    df = pd.DataFrame(bank)
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="NganHangMau")
    return out.getvalue()

